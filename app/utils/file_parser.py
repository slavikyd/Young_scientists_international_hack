import csv
import io
import logging
from typing import List, Dict, Any
import openpyxl

logger = logging.getLogger(__name__)


def parse_csv(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Parse CSV file content.
    
    Args:
        file_content: File content as bytes
        
    Returns:
        List of dictionaries with parsed data
    """
    try:
        # Decode bytes to string
        text = file_content.decode('utf-8')
        
        # Parse CSV
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        
        logger.info(f"Parsed {len(rows)} rows from CSV")
        return rows
        
    except Exception as e:
        logger.error(f"Error parsing CSV: {e}")
        raise


def parse_xlsx(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Parse XLSX file content.
    
    Args:
        file_content: File content as bytes
        
    Returns:
        List of dictionaries with parsed data
    """
    try:
        # Load workbook from bytes
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        # Get header row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Parse rows
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data[header] = cell_value
            
            # Only add non-empty rows
            if any(row_data.values()):
                rows.append(row_data)
        
        logger.info(f"Parsed {len(rows)} rows from XLSX")
        return rows
        
    except Exception as e:
        logger.error(f"Error parsing XLSX: {e}")
        raise